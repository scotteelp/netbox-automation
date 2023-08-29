# Logging Module
import os
import sys
import logging
import importlib

def setup_logging():
    logger = logging.getLogger(__name__)
    logger.setLevel(logging.DEBUG)

    # Create a formatter
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

    # Create a handler for writing log messages to a file
    file_handler = logging.FileHandler("netbox_api.log")
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    # Create a handler for displaying log messages in the terminal
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.WARNING)  # Only show warnings and errors in terminal
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)

    return logger

logger = setup_logging()

# Validate config.py
def validate_config():
    try:
        # Try importing NETBOX_TOKEN and NETBOX_URL from config.py
        from config import NETBOX_TOKEN, NETBOX_URL
        # Set environment variables
        os.environ["NETBOX_TOKEN"] = NETBOX_TOKEN
        os.environ["NETBOX_URL"] = NETBOX_URL
        print(BOLD + BG_CYAN + WHITE + "✅  Configuration validated successfully." + RESET)
        logger.info("✅  Configuration validated successfully.")
    except ImportError:
        logger.error("Configuration error: Missing or incomplete data in config.py.")
        print("Configuration error: Missing or incomplete data in config.py.")
        logger.error("Please provide valid NETBOX_URL and NETBOX_TOKEN.")
        print("Please provide valid NETBOX_URL and NETBOX_TOKEN.")
        sys.exit(1)

# Module Validation
import subprocess

# Import color constants from color_definitions.py
from color_definitions import BOLD, UNDERLINE, RESET, RED, GREEN, YELLOW, BLACK, BLUE, MAGENTA, CYAN, WHITE, BG_BLACK, BG_RED, BG_GREEN, BG_YELLOW, BG_BLUE, BG_MAGENTA, BG_CYAN, BG_WHITE

# List of required modules
required_modules = ['pynetbox', 'csv', 'sys', 'requests', 'pandas', 'datetime', 'openpyxl']

# Check if a module is installed
def is_module_installed(module_name):
    try:
        importlib.import_module(module_name)
        return True
    except ImportError:
        warning_message = f"❌  Module {module_name} is not installed."
        logger.warning(warning_message)
        #logger.info(warning_message)  # Log the same message as info level
        return False

# Install a missing module		
def install_module(module_name):
    try:
        # Redirect pip3 output to the log file
        with open("netbox_api.log", "a") as log_file:
            subprocess.check_call(["pip3", "install", module_name], stdout=log_file, stderr=log_file)

        success_message = f"Successfully installed {module_name}"
        print(BOLD + BG_CYAN + WHITE + f"✅  {success_message}" + RESET)
        logger.info(success_message)
        logger.debug(success_message)  # Log the same message as debug level
    except Exception as e:
        error_message = f"❌  Error installing {module_name}: {e}"
        print(BOLD + UNDERLINE + RED + f"❌  {error_message}" + RESET)
        logger.error(error_message)
        logger.debug(error_message)  # Log the same message as debug level

# Check and install missing modules
def check_and_install_modules(module_list):
    missing_modules = [module for module in module_list if not is_module_installed(module)]
    
    if missing_modules:
        #print("⚠️  The following required modules are missing:")
        #logger.warning("The following required modules are missing:")
        for module in missing_modules:
            #print(f" - {module}")
            logger.warning(f" - {module}")
            #logger.info(f" - {module}")
        
        while True:
            install_choice = input(BOLD + WHITE + "Do you want to install the missing modules? (y/n): " + RESET).strip().lower()

            if install_choice == "y" or install_choice == "yes":
                logger.info(BOLD + WHITE + "User chose to install missing modules.")
                for module in missing_modules:
                    install_module(module)
                break
            elif install_choice == "":
                print(BG_BLUE + BOLD + YELLOW + "Please enter 'y' or 'yes' to install missing modules." + RESET)
                logger.info("User pressed Enter/Return. Please enter 'y' or 'yes' to install missing modules." + RESET)
            else:
                print(BG_BLUE + BOLD + YELLOW + "Missing modules will not be installed." + RESET)
                logger.info("User chose not to install missing modules.")
                break
    
    else:
        print(BOLD + BG_CYAN + WHITE + "✅  All required modules are already installed." + RESET)
        logger.info("✅  All required modules are already installed.")
        logger.debug("✅  All required modules are already installed.")  # Log the same message as debug level

# Main Modules
import os
import pynetbox
import json
import csv
import sys
from csv import writer
import random
import requests
import pandas as pd
import datetime
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
import textwrap
import subprocess

# Load sensitive data from config.py and store as environment variables
try:
    from config import NETBOX_TOKEN, NETBOX_URL
    os.environ["NETBOX_TOKEN"] = NETBOX_TOKEN
    os.environ["NETBOX_URL"] = NETBOX_URL
except ImportError:
    print(BG_RED + BLACK + "Error: The config.py file is missing or incomplete." + RESET)
    sys.exit(1)

# Import ascii art from ascii_art.py
from ascii_art import VIDGO_ASCII, FACE_ASCII, CHUCK_ASCII, NETBOX_ASCII
		
#host and token
#nb = pynetbox.api(NETBOX_URL, NETBOX_TOKEN)
#fetch all devices
#nb_devicelist = nb.dcim.devices.all()

def display_config_file():
    try:
        completed_process = subprocess.run(["cat", "config.py"], check=True, text=True, capture_output=True)
        output = completed_process.stdout
        colored_output = f"\033[1m\033[41m\033[33m{output}\033[0m"  # Set BOLD, background to red, and text color to yellow
        print(colored_output + RESET)

        # Log each line separately
        for line in output.splitlines():
            logger.info(line)

    except subprocess.CalledProcessError as e:
        error_message = f"Error running 'cat' command: {e}"
        print(error_message)
        logger.error(error_message)
    except FileNotFoundError:
        error_message = "'cat' command not found. Make sure you're running on a Unix-like system."
        print(error_message)
        logger.error(error_message)
    except Exception as e:
        error_message = f"An error occurred: {e}"
        print(error_message)
        logger.error(error_message)
        
try:
    # Initialize the NetBox API connection
    nb = pynetbox.api(NETBOX_URL, NETBOX_TOKEN)
except IndexError as index_error:
    error_message = "Error initializing NetBox API: {}".format(index_error)
    logger.error(error_message)
    print(error_message)
    logger.error("Check your config.py file for `NETBOX_TOKEN` and `NETBOX_URL` definitions.")
    print(BOLD + BG_RED + YELLOW + "Check your config.py file for `NETBOX_TOKEN` and `NETBOX_URL` definitions." + RESET)
    display_config_file()
    #logger.info(display_config_file())
    print()
    sys.exit(1)
except Exception as e:
    error_message = "An error occurred: {}".format(e)
    logger.error(error_message)
    print("An error occurred while initializing NetBox API.")
    sys.exit(1)

nb_devicelist = nb.dcim.devices.all()

headers = ['Name', 'Status', 'Site', 'Rack', 'Role', 'Manufacturer', 'Type', 'Owner', 'Birthday', 'Age (Months)', 'Service Contract', 'Warranty', 'Serial Number', 'Platform', 'Software', 'SW_Version', 'Primary IP']

def calculate_age_in_months(birthday):
	today = datetime.today()
	birth_date = datetime.strptime(birthday, '%Y-%m-%d')
	age_months = (today.year - birth_date.year) * 12 + today.month - birth_date.month
	return age_months
	
def csv_to_xlsx(headers, devices_data):
	wb = openpyxl.Workbook()
	ws = wb.active

	ws.append(headers)
	for device in devices_data:
		ws.append([device.get(header, '') for header in headers])

	for col in ws.columns:
		max_length = 0
		column = col[0].column_letter  # Get the column name
		for cell in col:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(cell.value)
			except:
				pass
		adjusted_width = (max_length + 2)
		ws.column_dimensions[column].width = adjusted_width

	for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
		for cell in row:
			cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

	wb.save('output.xlsx')
	
def get_devices(nb_devicelist, headers):
    devices_data = []  # List to hold device information
    logger.info("Getting device information from Netbox...")
    print()
    print(BOLD + BG_GREEN + WHITE +"Getting device information from Netbox..." + RESET)
    print(UNDERLINE + BG_GREEN + BLACK + "................................................" + RESET)
    print()
    print(GREEN + NETBOX_ASCII + RESET)
    str1 = 'Active'
    for nb_device in nb_devicelist:
        result = {}
        status = str(nb_device.status)
        if status == str1:
            result['Name'] = str(nb_device)
            result['Status'] = status
            result['Site'] = str(nb_device.site)
            result['Rack'] = str(nb_device.rack)
            result['Role'] = nb_device.device_role.name
            result['Manufacturer'] = nb_device.device_type.manufacturer.name
            result['Type'] = str(nb_device.device_type)
            result['Owner'] = nb_device.custom_fields.get('owner')
            result['Birthday'] = nb_device.custom_fields.get('Birthday')
            age = nb_device.custom_fields.get('age')
            if age is None and result['Birthday']:
                result['Age (Months)'] = calculate_age_in_months(result['Birthday'])
            else:
                result['Age (Months)'] = age
            result['Service Contract'] = nb_device.custom_fields.get('service_contract')
            result['Warranty'] = nb_device.custom_fields.get('warranty')
            result['Serial Number'] = str(nb_device.serial)
            result['Platform'] = str(nb_device.platform)
            result['SW'] = nb_device.custom_fields.get('SW')
            result['SW_Version'] = nb_device.custom_fields.get('SW_Version')
            result['Primary IP'] = str(nb_device.primary_ip)

            devices_data.append(result)

        # Logging information for each device processed
        if 'Name' in result:
            logger.info("Processed device: %s", result['Name'])

        with open('output.csv', 'a', newline='') as f_object:
            writer_object = writer(f_object)
            for device in devices_data:
                writer_object.writerow([device.get(header, '') for header in headers])

    csv_to_xlsx(headers, devices_data)
    logger.info("Device information written to output.csv and output.xlsx")
    print(BOLD + BG_GREEN + WHITE + "Device information written to output.csv and output.xlsx" + RESET)
    logger.info("Finished getting device information from Netbox")
    print(BOLD + BG_GREEN + WHITE + "Finished getting device information from Netbox" + RESET)
    print(UNDERLINE + BG_GREEN + BLACK + "................................................" + RESET)
    print()


def update_age(nb_devicelist):
    logger.info("Updating age information for devices...")
    print()
    print(BG_CYAN + BLACK + "Updating age information for devices..." + RESET)
    print(UNDERLINE + BG_CYAN + BLACK + "................................................" + RESET)
    print()
    print(CYAN + NETBOX_ASCII + RESET)
    for nb_device in nb_devicelist:
        status = str(nb_device.status)
        if status == "Active":
            birthday = nb_device.custom_fields.get('Birthday')
            if birthday:
                new_age = calculate_age_in_months(birthday)
                # Update the 'age' custom field in NetBox
                nb_device.custom_fields['age'] = new_age
                nb_device.save()
                logger.info("Updated age for device %s to %d months.", nb_device.name, new_age)
    
    logger.info("Age information update complete.")
    print()
    print(UNDERLINE + BG_CYAN + BLACK + "................................................" + RESET)
    print()


def save_rack_details_to_xlsx(racks_with_devices):
    wb = openpyxl.Workbook()
    # Remove the default "Sheet"
    wb.remove(wb.active)

    for rack_name, devices_info in racks_with_devices.items():
        ws = wb.create_sheet(title=rack_name)
        ws.append(["Device Name", "Role", "Type", "Manufacturer", "Rack Unit"])
        for device_info in devices_info:
            ws.append([
                device_info["name"],
                device_info["role"],
                device_info["type"],
                device_info["manufacturer"],
                device_info["rack_unit"]
            ])

    wb.save('rack_details_with_devices.xlsx')

def get_rack_details_with_devices(nb_instance):
    try:
        logger.info("Fetching rack details and associated devices from NetBox...")
        print(BOLD + BG_GREEN + WHITE + "Fetching rack details and associated devices from NetBox..." + RESET)
        print(UNDERLINE + BG_GREEN + BLACK + "................................................" + RESET)

        print(GREEN + NETBOX_ASCII + RESET)
        racks_with_devices = {}

        # Fetch all racks from NetBox
        racks = nb_instance.dcim.racks.all()

        for rack in racks:
            rack_info = {
                "name": rack.name,
                "site": rack.site.name,
                "location": rack.location,
                "height": rack.u_height
            }

            devices_info = []
            for device in nb_instance.dcim.devices.filter(rack_id=rack.id):
                device_info = {
                    "name": device.name,
                    "role": device.device_role.name if device.device_role else "N/A",
                    "type": device.device_type.model if device.device_type else "N/A",
                    "manufacturer": device.device_type.manufacturer.name if device.device_type and device.device_type.manufacturer else "N/A",
                    "rack_unit": device.position if device.position else "N/A"
                }
                devices_info.append(device_info)

            racks_with_devices[rack.name] = devices_info

        logger.info("Retrieved rack details and associated devices.")
        print(BOLD + BG_GREEN + WHITE + "Retrieved rack details and associated devices." + RESET)
        save_rack_details_to_xlsx(racks_with_devices)
        logger.info("Saved rack details with associated devices to rack_details_with_devices.xlsx")
        print(BOLD + BG_GREEN + WHITE + "Saved rack details with associated devices to rack_details_with_devices.xlsx" + RESET)

    except pynetbox.RequestError as pnb_error:
        logger.error("A pynetbox error occurred: %s", pnb_error)
    except Exception as e:
        logger.error("An error occurred: %s", e)
    print(UNDERLINE + BG_GREEN + BLACK + "................................................" + RESET)
    print()


def get_rack_names(nb_instance):
    rack_names = []
    racks = nb_instance.dcim.racks.all()
    for rack in racks:
        rack_names.append(rack.name)
    return rack_names

def joke():
    try:
        response = requests.get("https://api.chucknorris.io/jokes/random")
        response.raise_for_status()  # Check for HTTP errors
        joke_data = response.json()
        joke_text = joke_data.get("value")
        logger.info("Random Chuck Norris Joke: %s", joke_text)  # Log the full joke text

        print()  # Print an empty line
        print(YELLOW + CHUCK_ASCII + RESET)
        print()

        # Word wrap the joke text and print it with background color reset after 77 characters
        wrapped_joke_text = textwrap.fill(joke_text, width=62)  # Adjust width as needed
        lines = wrapped_joke_text.split('\n')
        for line in lines:
            print(BG_YELLOW + BLACK + line[:62] + RESET)
            print(BG_YELLOW + BLACK + line[62:] + RESET)
        print("................................................................")
        print("________________________________________________________________")    

    except requests.exceptions.ConnectionError:
        logger.error("Error: Could not establish an internet connection.")
    except requests.exceptions.HTTPError as e:
        logger.error("HTTP error occurred: %s", e)
    except Exception as e:
        logger.error("An error occurred: %s", e)

def show_help():
    print()
    print(RED + VIDGO_ASCII + RESET)
    print(RED + FACE_ASCII + RESET)
    print(BOLD + "Available functions:" + RESET)
    print(" ► " + BG_GREEN + BLACK + "get_devices" + RESET + " or " + BG_GREEN + BLACK + "-d" + RESET + " ► GETS active device info from Netbox, writes output.csv and converts to output.xlsx file.")
    print (" ► " + BG_GREEN + BLACK + "get_racks" + RESET + " or " + BG_GREEN + BLACK + "-r" + RESET + " ► GETS rack with device details, and saves file rack_details_with_devices.xlsx.")
    print(" ► " + BG_BLUE + WHITE + "update_age" + RESET + " or " + BG_BLUE + WHITE + "-a" + RESET + " ► This will update the age for all active devices on Netbox server.")
    print(" ► " + BG_YELLOW + BLACK + "joke:" + RESET + " or " + BG_YELLOW + BLACK + "-j" + RESET +  " ► Prints random Chuck Norris joke.")
    print(" ► " + BG_WHITE + BLACK + "validate_config:" + RESET + " or " + BG_WHITE + BLACK + "-v" + RESET +  " ► Validates script config.py file.")
    print(BOLD + WHITE + " ► Usage: python netbox_api.py <function_name>")
    print(UNDERLINE + BG_CYAN + "................................................" + RESET)
    print()
    
    # Log the help information
    logger.info("Displayed help information to terminal.")
    
    sys.exit(0)

	
def main():
    try:
        # Initialize logger
        logger.info("Script started")
        
        # Validate config.py
        validate_config()

        # Set up NetBox API connection
        nb = pynetbox.api(NETBOX_URL, NETBOX_TOKEN)

        # Get list of available rack names
        rack_names = get_rack_names(nb)

        # Call the function to check and install modules
        logger.setLevel(logging.ERROR)
        check_and_install_modules(required_modules)
        #logger.setLevel(logging.DEBUG)
        
        if len(sys.argv) < 2:
            show_help()

        function_name = sys.argv[1]

        if function_name == "get_devices" or function_name == "-d":
            get_devices(nb_devicelist, headers)
        elif function_name == "update_age" or function_name == "-a":
            update_age(nb_devicelist)
        elif function_name == "get_racks" or function_name == "-r":
             get_rack_details_with_devices(nb)
        elif function_name == "joke" or function_name == "-j":
            joke()
        elif function_name == "validate_config" or function_name == "-v":
            validate_config()
        elif function_name == "--help" or function_name == "-h":
            show_help()
        else:
            logger.error(f"Function '{function_name}' not recognized.")

    except pynetbox.RequestError as pnb_error:
        logger.error("A pynetbox error occurred: %s", pnb_error)
    except Exception as e:
        logger.error(f"An error occurred: {e}")
    
    finally:
        # Log a message at the end of the script run
        logger.info("Script completed")

if __name__ == "__main__":
    main()
